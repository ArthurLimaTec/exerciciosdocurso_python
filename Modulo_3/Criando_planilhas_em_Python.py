# Crie uma planilha Excel geral de funcionários de uma empresa 
# Com ID, nome, departamento, telefone e e-mail.
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from faker import Faker
import random

# Caminho de destino do arquivo:
ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'Planilha_de_Funcionarios.xlsx'

workbook = Workbook()

# Nomeando a planilha:
sheet_name = 'Funcionarios'

# Criamos a planilha
workbook.create_sheet(sheet_name, 0)

# Selecionando a planilha:
worksheet: Worksheet = workbook[sheet_name]

# Criando os cabeçalhos:
worksheet.cell(1, 1, 'ID')
worksheet.cell(1, 2, 'Nome')
worksheet.cell(1, 3, 'Departamento')
worksheet.cell(1, 4, 'Telefone')
worksheet.cell(1, 5, 'E-mail')

##############################################################################################

# Inicializando o Faker para gerar dados fictícios
fake = Faker()

# Lista de departamentos fictícios e prefixos:
departamentos = ['Vendas', 'Marketing', 'RH', 'TI', 'Produção']
telefones_prefixo = ['(11)', '(13)', '(21)', '(41)', '(51)', '(61)', '(71)', '(81)', '(91)']

# Função para gerar telefones aleatórios:
def gerar_telefone():
    prefixo = random.choice(telefones_prefixo)
    numero = ''.join(random.choices('0123456789', k=8))
    return f"{prefixo} {numero[:4]}-{numero[4:]}"

# Lista para armazenar os dados:
dados = []

# Gerando 100 linhas de dados aleatórios
for i in range(1, 101):
    id_ = i
    nome = fake.name()
    departamento = random.choice(departamentos)
    telefone = gerar_telefone()
    email = fake.email()
    dados.append([id_, nome, departamento, telefone, email])


for funcionario in dados:
    worksheet.append(funcionario)

workbook.save(WORKBOOK_PATH)

